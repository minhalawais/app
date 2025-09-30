from flask import jsonify, request, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt, get_jwt_identity
from . import main
from ..crud import customer_crud
from werkzeug.utils import secure_filename
import os
import tempfile
import csv
import io
import uuid
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import json

UPLOAD_FOLDER = os.path.join(current_app.root_path, 'uploads/cnic_images')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))

async def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# New route for handling immediate file uploads
@main.route('/customers/upload-file/<string:file_type>', methods=['POST'])
@jwt_required()
async def upload_customer_file(file_type):
    claims = get_jwt()
    company_id = claims['company_id']
    
    if file_type not in ['cnic_front_image', 'cnic_back_image', 'agreement_document']:
        return jsonify({'error': 'Invalid file type'}), 400
    
    if file_type not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files[file_type]
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        # Generate a unique filename with UUID to prevent collisions
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4()}_{file_type}.{file_extension}"
        
        # Create relative path (this is important)
        relative_path = os.path.join('uploads/cnic_images', unique_filename)
        file_path = os.path.join(PROJECT_ROOT, relative_path)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Save the file
        file.save(file_path)
        
        # Return the relative file path to be stored in the customer record
        return jsonify({
            'success': True,
            'file_path': relative_path,  # Return relative path
            'file_name': unique_filename,
            'file_type': file_extension,
            'message': 'File uploaded successfully'
        }), 200
    
    return jsonify({'error': 'Invalid file format'}), 400

@main.route('/customers/list', methods=['GET'])
@jwt_required()
async def get_customers():
    claims = get_jwt()
    company_id = claims['company_id']
    user_role = claims['role']
    employee_id = get_jwt_identity()
    customers = await customer_crud.get_all_customers(company_id, user_role, employee_id)
    return jsonify(customers), 200

@main.route('/customers/check-internet-id/<string:internet_id>', methods=['GET'])
@jwt_required()
def check_internet_id_availability(internet_id):
    claims = get_jwt()
    company_id = claims['company_id']
    
    customer_id = request.args.get('customer_id')
    
    existing_customer = customer_crud.check_existing_internet_id(internet_id, company_id)
    
    if existing_customer and customer_id and str(existing_customer.id) == customer_id:
        return jsonify({
            'available': True,
            'message': 'Current Internet ID'
        }), 200
    
    return jsonify({
        'available': existing_customer is None,
        'message': 'Internet ID is available' if existing_customer is None else 'Internet ID already exists'
    }), 200

@main.route('/customers/check-cnic/<string:cnic>', methods=['GET'])
@jwt_required()
def check_cnic_availability(cnic):
    claims = get_jwt()
    company_id = claims['company_id']
    
    customer_id = request.args.get('customer_id')
    
    existing_customer = customer_crud.check_existing_cnic(cnic, company_id)
    
    if existing_customer and customer_id and str(existing_customer.id) == customer_id:
        return jsonify({
            'available': True,
            'message': 'Current CNIC'
        }), 200
    
    return jsonify({
        'available': existing_customer is None,
        'message': 'CNIC is available' if existing_customer is None else 'CNIC already exists'
    }), 200

@main.route('/customers/add', methods=['POST'])
@jwt_required()
async def add_new_customer():
    claims = get_jwt()
    company_id = claims['company_id']
    user_role = claims['role']
    current_user_id = get_jwt_identity()
    ip_address = request.remote_addr
    user_agent = request.headers.get('User-Agent')
    
    data = request.form.to_dict()
    data['company_id'] = company_id
    
    # Validate data first
    validation_errors = await customer_crud.validate_customer_data(data, is_update=False)
    
    if validation_errors:
        return jsonify({'errors': validation_errors}), 400
    
    try:
        # Check Internet ID uniqueness
        existing_internet_id = customer_crud.check_existing_internet_id(data['internet_id'], company_id)
        if existing_internet_id:
            return jsonify({'errors': {'internet_id': 'Internet ID already exists'}}), 400
        
        # Check CNIC uniqueness
        existing_cnic = customer_crud.check_existing_cnic(data['cnic'], company_id)
        if existing_cnic:
            return jsonify({'errors': {'cnic': 'CNIC already exists'}}), 400
        
        new_customer = await customer_crud.add_customer(data, user_role, current_user_id, ip_address, user_agent, company_id)
        return jsonify({'message': 'Customer added successfully', 'id': str(new_customer.id)}), 201
        
    except ValueError as ve:
        return jsonify({'error': 'Validation failed', 'message': str(ve)}), 400
    except Exception as e:
        logger.error(f"Error adding customer: {str(e)}")
        return jsonify({'error': 'Failed to add customer', 'message': 'An unexpected error occurred'}), 500

@main.route('/customers/update/<string:id>', methods=['PUT'])
@jwt_required()
async def update_existing_customer(id):
    claims = get_jwt()
    company_id = claims['company_id']
    user_role = claims['role']
    current_user_id = get_jwt_identity()
    ip_address = request.remote_addr
    user_agent = request.headers.get('User-Agent')
    data = request.form.to_dict()
    data['company_id'] = company_id
    
    # Validate data first
    validation_errors = await customer_crud.validate_customer_data(data, is_update=True, customer_id=id)
    
    if validation_errors:
        return jsonify({'errors': validation_errors}), 400
    
    try:
        # Check CNIC uniqueness (excluding current customer)
        if data.get('cnic'):
            existing_cnic_customer = customer_crud.check_existing_cnic(data['cnic'], company_id)
            if existing_cnic_customer and str(existing_cnic_customer.id) != id:
                return jsonify({'errors': {'cnic': 'CNIC already exists'}}), 400
        
        updated_customer = await customer_crud.update_customer(id, data, company_id, user_role, current_user_id, ip_address, user_agent)
        if updated_customer:
            return jsonify({'message': 'Customer updated successfully'}), 200
        return jsonify({'message': 'Customer not found'}), 404
        
    except ValueError as ve:
        return jsonify({'error': 'Validation failed', 'message': str(ve)}), 400
    except Exception as e:
        logger.error(f"Error updating customer: {str(e)}")
        return jsonify({'error': 'Failed to update customer', 'message': 'An unexpected error occurred'}), 500

@main.route('/customers/delete/<string:id>', methods=['DELETE'])
@jwt_required()
async def delete_existing_customer(id):
    claims = get_jwt()
    company_id = claims['company_id']
    user_role = claims['role']
    current_user_id = get_jwt_identity()
    ip_address = request.remote_addr
    user_agent = request.headers.get('User-Agent')
    if await customer_crud.delete_customer(id, company_id, user_role, current_user_id, ip_address, user_agent):
        return jsonify({'message': 'Customer deleted successfully'}), 200
    return jsonify({'message': 'Customer not found'}), 404

@main.route('/customers/toggle-status/<string:id>', methods=['PATCH'])
@jwt_required()
async def toggle_customer_active_status(id):
    claims = get_jwt()
    company_id = claims['company_id']
    user_role = claims['role']
    current_user_id = get_jwt_identity()
    ip_address = request.remote_addr
    user_agent = request.headers.get('User-Agent')
    customer = await customer_crud.toggle_customer_status(id, company_id, user_role, current_user_id, ip_address, user_agent)
    if customer:
        return jsonify({'message': f"Customer {'activated' if customer.is_active else 'deactivated'} successfully"}), 200
    return jsonify({'message': 'Customer not found'}), 404

@main.route('/customers/<string:id>', methods=['GET'])
@jwt_required()
async def get_customer_details(id):
    claims = get_jwt()
    company_id = claims['company_id']
    customer = await customer_crud.get_customer_details(id, company_id)
    if customer:
        return jsonify(customer), 200
    return jsonify({'message': 'Customer not found'}), 404

@main.route('/invoices/customer/<string:id>', methods=['GET'])
@jwt_required()
async def get_customer_invoices(id):
    claims = get_jwt()
    company_id = claims['company_id']
    invoices = await customer_crud.get_customer_invoices(id, company_id)
    return jsonify(invoices), 200

@main.route('/payments/customer/<string:id>', methods=['GET'])
@jwt_required()
async def get_customer_payments(id):
    claims = get_jwt()
    company_id = claims['company_id']
    payments = await customer_crud.get_customer_payments(id, company_id)
    return jsonify(payments), 200

@main.route('/complaints/customer/<string:id>', methods=['GET'])
@jwt_required()
async def get_customer_complaints(id):
    claims = get_jwt()
    company_id = claims['company_id']
    complaints = await customer_crud.get_customer_complaints(id, company_id)
    return jsonify(complaints), 200

@main.route('/customers/cnic-front-image/<string:id>', methods=['GET'])
@jwt_required()
async def get_cnic_front_image(id):
    claims = get_jwt()
    company_id = claims['company_id']
    customer = await customer_crud.get_customer_cnic(id, company_id)
    if customer and customer.cnic_front_image:
        cnic_image_path = os.path.join(PROJECT_ROOT, customer.cnic_front_image)
        if os.path.exists(cnic_image_path):
            return send_file(cnic_image_path, mimetype='image/jpeg')
        else:
            return jsonify({'error': 'CNIC front image file not found'}), 404
    return jsonify({'error': 'CNIC front image not found'}), 404

@main.route('/customers/cnic-back-image/<string:id>', methods=['GET'])
@jwt_required()
async def get_cnic_back_image(id):
    claims = get_jwt()
    company_id = claims['company_id']
    customer = await customer_crud.get_customer_cnic(id, company_id)
    if customer and customer.cnic_back_image:
        cnic_image_path = os.path.join(PROJECT_ROOT, customer.cnic_back_image)
        if os.path.exists(cnic_image_path):
            print('Document : ', cnic_image_path)
            return send_file(cnic_image_path, mimetype='image/jpeg')
        else:
            return jsonify({'error': 'CNIC back image file not found'}), 404
    return jsonify({'error': 'CNIC back image not found'}), 404

@main.route('/customers/agreement-document/<string:id>', methods=['GET'])
@jwt_required()
async def get_agreement_document(id):
    claims = get_jwt()
    company_id = claims['company_id']
    customer = await customer_crud.get_customer_details(id, company_id)
    if customer and customer['agreement_document']:
        agreement_document_path = os.path.join(PROJECT_ROOT, customer['agreement_document'])
        if os.path.exists(agreement_document_path):
            print('Document : ', agreement_document_path)
            return send_file(agreement_document_path, mimetype='image/jpeg')
        else:
            return jsonify({'error': 'Agreement document file not found'}), 404
    return jsonify({'error': 'Agreement document not found'}), 404

@main.route('/customers/template', methods=['GET'])
@jwt_required()
async def get_customer_template():
    """Generate and return an Excel template with dropdowns and validation for bulk customer import"""
    claims = get_jwt()
    company_id = claims['company_id']
    
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Import Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    required_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Define headers with validation info
    headers = [
        {'name': 'internet_id', 'required': True, 'comment': 'Unique Internet ID (e.g., NET12345)'},
        {'name': 'first_name', 'required': True, 'comment': 'Customer first name'},
        {'name': 'last_name', 'required': True, 'comment': 'Customer last name'},
        {'name': 'email', 'required': True, 'comment': 'Valid email address'},
        {'name': 'phone_1', 'required': True, 'comment': 'Primary phone number (with country code)'},
        {'name': 'phone_2', 'required': False, 'comment': 'Secondary phone number (optional)'},
        {'name': 'area_id', 'required': True, 'comment': 'Select from dropdown list'},
        {'name': 'installation_address', 'required': True, 'comment': 'Complete installation address'},
        {'name': 'service_plan_id', 'required': True, 'comment': 'Select from dropdown list'},
        {'name': 'isp_id', 'required': True, 'comment': 'Select from dropdown list'},
        {'name': 'connection_type', 'required': True, 'comment': 'internet, tv_cable, or both'},
        {'name': 'internet_connection_type', 'required': False, 'comment': 'wire or wireless (required if connection_type is internet/both)'},
        {'name': 'tv_cable_connection_type', 'required': False, 'comment': 'analog or digital (required if connection_type is tv_cable/both)'},
        {'name': 'installation_date', 'required': True, 'comment': 'Format: YYYY-MM-DD'},
        {'name': 'cnic', 'required': True, 'comment': '13-digit CNIC number'},
        {'name': 'gps_coordinates', 'required': False, 'comment': 'Format: latitude,longitude (e.g., 31.5204,74.3587)'}
    ]
    
    # Set column widths and create headers
    for col_idx, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
        
        # Set header cell
        cell = ws.cell(row=1, column=col_idx, value=header['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Add comment with instructions
        cell.comment = Comment(header['comment'], "System")
        
        # Color code required vs optional fields in row 2
        info_cell = ws.cell(row=2, column=col_idx, 
                           value="REQUIRED" if header['required'] else "OPTIONAL")
        info_cell.fill = required_fill if header['required'] else optional_fill
        info_cell.font = Font(bold=True, size=8)
        info_cell.alignment = Alignment(horizontal='center')
        info_cell.border = border
    
    # Fetch dropdown data from database
    try:
        # Get areas, service plans, and ISPs for dropdowns
        areas = await customer_crud.get_company_areas(company_id)
        service_plans = await customer_crud.get_company_service_plans(company_id)
        isps = await customer_crud.get_company_isps(company_id)
        
        # Create hidden sheets for dropdown data
        area_sheet = wb.create_sheet("Areas")
        plan_sheet = wb.create_sheet("ServicePlans")
        isp_sheet = wb.create_sheet("ISPs")
        
        # Populate area data
        area_sheet.cell(row=1, column=1, value="ID")
        area_sheet.cell(row=1, column=2, value="Name")
        for idx, area in enumerate(areas, 2):
            area_sheet.cell(row=idx, column=1, value=str(area['id']))
            area_sheet.cell(row=idx, column=2, value=area['name'])
        
        # Populate service plan data
        plan_sheet.cell(row=1, column=1, value="ID")
        plan_sheet.cell(row=1, column=2, value="Name")
        for idx, plan in enumerate(service_plans, 2):
            plan_sheet.cell(row=idx, column=1, value=str(plan['id']))
            plan_sheet.cell(row=idx, column=2, value=plan['name'])
        
        # Populate ISP data
        isp_sheet.cell(row=1, column=1, value="ID")
        isp_sheet.cell(row=1, column=2, value="Name")
        for idx, isp in enumerate(isps, 2):
            isp_sheet.cell(row=idx, column=1, value=str(isp['id']))
            isp_sheet.cell(row=idx, column=2, value=isp['name'])
        
        # Hide the data sheets
        area_sheet.sheet_state = 'hidden'
        plan_sheet.sheet_state = 'hidden'
        isp_sheet.sheet_state = 'hidden'
        
        # Add data validation for dropdowns
        # Area dropdown (column 7)
        area_validation = DataValidation(
            type="list",
            formula1=f"Areas!$A$2:$A${len(areas)+1}",
            showDropDown=True
        )
        area_validation.error = "Please select a valid area from the dropdown"
        area_validation.errorTitle = "Invalid Area"
        ws.add_data_validation(area_validation)
        area_validation.add(f"G3:G1000")  # Apply to area_id column
        
        # Service Plan dropdown (column 9)
        plan_validation = DataValidation(
            type="list",
            formula1=f"ServicePlans!$A$2:$A${len(service_plans)+1}",
            showDropDown=True
        )
        plan_validation.error = "Please select a valid service plan from the dropdown"
        plan_validation.errorTitle = "Invalid Service Plan"
        ws.add_data_validation(plan_validation)
        plan_validation.add(f"I3:I1000")  # Apply to service_plan_id column
        
        # ISP dropdown (column 10)
        isp_validation = DataValidation(
            type="list",
            formula1=f"ISPs!$A$2:$A${len(isps)+1}",
            showDropDown=True
        )
        isp_validation.error = "Please select a valid ISP from the dropdown"
        isp_validation.errorTitle = "Invalid ISP"
        ws.add_data_validation(isp_validation)
        isp_validation.add(f"J3:J1000")  # Apply to isp_id column
        
    except Exception as e:
        print(f"Error fetching dropdown data: {e}")
    
    # Add validation for other fields
    # Connection type validation
    connection_validation = DataValidation(
        type="list",
        formula1='"internet,tv_cable,both"',
        showDropDown=True
    )
    connection_validation.error = "Please select: internet, tv_cable, or both"
    connection_validation.errorTitle = "Invalid Connection Type"
    ws.add_data_validation(connection_validation)
    connection_validation.add("K3:K1000")  # Apply to connection_type column
    
    # Internet connection type validation
    internet_conn_validation = DataValidation(
        type="list",
        formula1='"wire,wireless"',
        showDropDown=True
    )
    internet_conn_validation.error = "Please select: wire or wireless"
    internet_conn_validation.errorTitle = "Invalid Internet Connection Type"
    ws.add_data_validation(internet_conn_validation)
    internet_conn_validation.add("L3:L1000")  # Apply to internet_connection_type column
    
    # TV cable connection type validation
    tv_conn_validation = DataValidation(
        type="list",
        formula1='"analog,digital"',
        showDropDown=True
    )
    tv_conn_validation.error = "Please select: analog or digital"
    tv_conn_validation.errorTitle = "Invalid TV Cable Connection Type"
    ws.add_data_validation(tv_conn_validation)
    tv_conn_validation.add("M3:M1000")  # Apply to tv_cable_connection_type column
    
    # Email validation
    email_validation = DataValidation(
        type="custom",
        formula1='ISERROR(FIND("@",E3))=FALSE',
        showDropDown=False
    )
    email_validation.error = "Please enter a valid email address"
    email_validation.errorTitle = "Invalid Email"
    ws.add_data_validation(email_validation)
    email_validation.add("E3:E1000")  # Apply to email column
    
    # CNIC validation (13 digits)
    cnic_validation = DataValidation(
        type="textLength",
        operator="equal",
        formula1="13",
        showDropDown=False
    )
    cnic_validation.error = "CNIC must be exactly 13 digits"
    cnic_validation.errorTitle = "Invalid CNIC"
    ws.add_data_validation(cnic_validation)
    cnic_validation.add("O3:O1000")  # Apply to cnic column
    
    # Add example row
    example_row = [
        'NET12345', 'John', 'Doe', 'john.doe@example.com', '923001234567', '923007654321',
        'area-uuid-here', '123 Main St, City', 'service-plan-uuid-here', 'isp-uuid-here',
        'internet', 'wire', '', '2023-05-01', '1234512345671', '31.5204,74.3587'
    ]
    
    for col_idx, value in enumerate(example_row, 1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.border = border
        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Add instructions sheet
    instructions_sheet = wb.create_sheet("Instructions", 0)
    instructions = [
        "CUSTOMER BULK IMPORT INSTRUCTIONS",
        "",
        "1. REQUIRED FIELDS (marked in red):",
        "   - All required fields must be filled",
        "   - Use the dropdown lists for area_id, service_plan_id, and isp_id",
        "",
        "2. FIELD FORMATS:",
        "   - internet_id: Unique identifier (e.g., NET12345)",
        "   - email: Valid email format (user@domain.com)",
        "   - phone_1/phone_2: Include country code (92XXXXXXXXX)",
        "   - installation_date: YYYY-MM-DD format",
        "   - cnic: Exactly 13 digits",
        "   - connection_type: Choose from internet, tv_cable, or both",
        "",
        "3. CONDITIONAL REQUIREMENTS:",
        "   - If connection_type is 'internet' or 'both', internet_connection_type is required",
        "   - If connection_type is 'tv_cable' or 'both', tv_cable_connection_type is required",
        "",
        "4. VALIDATION:",
        "   - The system will validate all data before import",
        "   - Errors will be shown with specific row and field information",
        "   - You can edit invalid rows directly in the validation interface",
        "",
        "5. TIPS:",
        "   - Use the example row as a reference",
        "   - Copy UUIDs from the dropdown sheets for area_id, service_plan_id, and isp_id",
        "   - Save the file before uploading"
    ]
    
    for idx, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=idx, column=1, value=instruction)
        if idx == 1:
            cell.font = Font(bold=True, size=16)
        elif instruction.endswith(":"):
            cell.font = Font(bold=True)
        instructions_sheet.column_dimensions['A'].width = 80
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # Return the file
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name='customer_import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@main.route('/customers/validate-bulk', methods=['POST'])
@jwt_required()
async def validate_bulk_customers():
    """Validate bulk customer data without saving to database"""
    claims = get_jwt()
    company_id = claims['company_id']
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Check file extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ['.csv', '.xls', '.xlsx']:
        return jsonify({'error': 'Invalid file format. Please upload a CSV or Excel file'}), 400
    
    # Save the file temporarily
    temp_file = tempfile.NamedTemporaryFile(delete=False)
    file.save(temp_file.name)
    temp_file.close()
    
    try:
        # Read the file based on its extension
        if file_ext == '.csv':
            df = pd.read_csv(temp_file.name)
        else:  # Excel file
            df = pd.read_excel(temp_file.name)
        print('Dataframe: ',df)
        # Validate the data without saving
        validation_results = await customer_crud.validate_bulk_customers(df, company_id)
        
        return jsonify(validation_results), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
    finally:
        # Clean up the temporary file
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

@main.route('/customers/bulk-add', methods=['POST'])
@jwt_required()
async def bulk_add_customers():
    """Process validated customer data and save to database"""
    claims = get_jwt()
    company_id = claims['company_id']
    user_role = claims['role']
    current_user_id = get_jwt_identity()
    ip_address = request.remote_addr
    user_agent = request.headers.get('User-Agent')
    
    if 'validatedData' in request.json:
        # Process pre-validated data
        validated_data = request.json['validatedData']
        results = await customer_crud.process_validated_customers(
            validated_data, 
            company_id, 
            user_role, 
            current_user_id, 
            ip_address, 
            user_agent
        )
        return jsonify(results), 200
    
    # Original file upload logic
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Check file extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ['.csv', '.xls', '.xlsx']:
        return jsonify({'error': 'Invalid file format. Please upload a CSV or Excel file'}), 400
    
    # Save the file temporarily
    temp_file = tempfile.NamedTemporaryFile(delete=False)
    file.save(temp_file.name)
    temp_file.close()
    
    try:
        # Read the file based on its extension
        if file_ext == '.csv':
            df = pd.read_csv(temp_file.name)
        else:  # Excel file
            df = pd.read_excel(temp_file.name)
        
        # Process the data
        results = await customer_crud.bulk_add_customers(
            df, 
            company_id, 
            user_role, 
            current_user_id, 
            ip_address, 
            user_agent
        )
        
        return jsonify(results), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
    finally:
        # Clean up the temporary file
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
